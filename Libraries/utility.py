from tkinter import simpledialog, messagebox, Radiobutton
import time
import tkinter as tk
from tkinter import ttk
import datetime
from openpyxl import load_workbook
import pandas as pd
import os
import re
import pyautogui
import pygetwindow as gw
from robot.api  import logger
import shutil
import openai
import json
import base64
import glob
import ast
import vertexai
import subprocess
from pathlib import Path

from vertexai.generative_models import (
    GenerativeModel,
    GenerationConfig,
    HarmCategory,
    HarmBlockThreshold,
    Part,
    SafetySetting
)

def show_message_popup(title, message):
    # Create a tkinter root window
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    # Show message popup
    messagebox.showinfo(title, message)

    # Bring the message popup window to the front
    root.attributes('-topmost', True)
    root.after_idle(root.attributes, '-topmost', False)  # Reset the topmost attribute after a short delay

def generate_current_date_string():
    # Get the current date and time
    current_datetime = datetime.datetime.now()

    # Extract minute, hour, year, month, and date components
    minute = current_datetime.minute
    hour = current_datetime.hour
    year = current_datetime.year % 100  # Taking only the last two digits of the year
    month = current_datetime.month
    date = current_datetime.day

    # Convert integers to strings and pad with zeros if necessary
    minute_str = str(minute).zfill(2)
    hour_str = str(hour).zfill(2)
    year_str = str(year).zfill(2)
    month_str = str(month).zfill(2)
    date_str = str(date).zfill(2)

    # Concatenate the components to form the date string
    date_string = minute_str + hour_str + year_str + month_str + date_str
    return date_string

# # Example usage:
# generated_date_string = generate_current_date_string()
# print("Generated date string:", generated_date_string)

def generate_dynamic_url(user_id, sid, base_url):
    # Format the URL with dynamic values for p_user_id and p_sid
    dynamic_url = f"{base_url}&p_user_id={user_id}&p_sid={sid}&p_pmc=1"
    return dynamic_url

# # Example usage:
# user_id = "elm.spr"
# sid = "5116240415"
# url = generate_dynamic_url(user_id, sid)
# print("Generated URL:", url)

def open_edge(url):
    # Close any running instances of Microsoft Edge
    os.system("taskkill /F /IM msedge.exe /T")
    # Wait for a moment to ensure the browser is closed
    time.sleep(2)
    pyautogui.press('alt')
    
    edge_path = "C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedge.exe"  # Path to Edge executable
    # webbrowser.register('edge', None, webbrowser.BackgroundBrowser(edge_path))
    # webbrowser.get('edge').open(url)
    subprocess.Popen([edge_path, "--start-maximized", url])
    
    
    
    
# # Example usage:
# open_edge("http://appsdms86.maruti.com:7779/forms/frmservlet?config=ggsmcnnoc1&p_user_id=elm.spr&p_sid=5116240415&p_pmc=1")

def check_sheet_exists(file_path, sheet_name):
    try:
        workbook = load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            print("File  found.")
            return True
            
        else:
            return False
            print("File not found.")
    except FileNotFoundError:
        print("File not found.")
        return False
    except Exception as e:
        print(f"An error occurred: {e}")
        return False
    
def login_read_credentials_from_excel(file_path):
    try:
        
        df = pd.read_excel(file_path, sheet_name='Sheet1')
        user_id = df.iloc[0, 0]  
        password = df.iloc[0, 1] 
        url = df.iloc[0, 2]  

        return user_id, password,url
    except Exception as e:
        print(f"An error occurred: {e}")
        return None, None
    
def get_location_value(file_path):
    # Read the Excel file
    df = pd.read_excel(file_path, sheet_name='Sheet1')
    
    # Get the value from the "Location" column
    location_value = df['Location'].iloc[0]
    
    return location_value

def login_with_alternative_url(excel_file_path):
    try:
        # Read the Excel file
        df = pd.read_excel(excel_file_path)

        # Extract the values assuming they are named 'userid', 'password', and 'alternative_url'
        userid = df['UserId'].iloc[0]
        password = df['Password'].iloc[0]
        alternative_url = df['Alternative_URL'].iloc[0]

        return userid, password, alternative_url

    except Exception as e:
        print(f"An error occurred: {e}")
        return None, None, None
    
def get_title_starting_with(window_info, prefix):
    """
    Function to get the title starting with the specified prefix from the window information dictionary.
    
    Args:
    - window_info (dict): Information dictionary of the window containing 'title' key.
    - prefix (str): Prefix string to match with the beginning of the title.

    Returns:
    - str or None: The title starting with the specified prefix, or None if not found.
    """
    title = window_info.get('title', '')
    if title.startswith(prefix):
        return title
    return None

def show_radio_message_box():
    root = tk.Tk()
    root.title("Please Select Your Option:")
    root.resizable(False, False)  # Disable resizing

    selected_option = tk.StringVar()

    # Function to get the selected option and close the window
    def get_selected_option():
        root.destroy()

    # Create a frame to contain the radio buttons
    frame = ttk.Frame(root)
    frame.pack(padx=10, pady=10)

    # Create radio buttons
    ttk.Radiobutton(frame, text="ALL", variable=selected_option, value="ALL").grid(row=0, column=0, padx=5, pady=5)
    ttk.Radiobutton(frame, text="ERP", variable=selected_option, value="ERP").grid(row=0, column=1, padx=5, pady=5)
    ttk.Radiobutton(frame, text="DMS", variable=selected_option, value="DMS").grid(row=0, column=2, padx=5, pady=5)

    # OK button to get the selected option
    ok_button = ttk.Button(root, text="OK", command=get_selected_option)
    ok_button.pack(pady=5)

    # Center the window
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry('{}x{}+{}+{}'.format(width, height, x, y))

    root.mainloop()

    return selected_option.get()

# Example usage
# selected_option = show_radio_message_box()
# print("Selected Option:", selected_option)

def open_application():
    # Path to the executable of the desktop application
    app_path = "C:\\Users\\Public\\Desktop\\Wings 23E Launcher.lnk"
    os.startfile(app_path)

def remove_unwanted_charecters(extracted_value):
    converted_value = str(extracted_value)
    
    cleaned_extracted_value = re.sub(r"[()\[\]{}]", "", converted_value)

    if "§" in cleaned_extracted_value:
        cleaned_extracted_value = cleaned_extracted_value.replace("§", "5")

    if "|" in cleaned_extracted_value:
        cleaned_extracted_value = cleaned_extracted_value.replace("|", "")

    print(cleaned_extracted_value)
    return  cleaned_extracted_value

def check_and_separate_yes_or_no(extracted_value):
    converted_value = str(extracted_value)
    if converted_value.startswith("Y"):
        return  "Y"
    else:
        return  "N"
    
def remove_space(extracted_value):
    if len(extracted_value) < 2:
        return extracted_value  # Return as is if length is less than 2
    
    converted_value = str(extracted_value)
    converted_value = converted_value.replace(" ", "")   
    
    return converted_value


def close_recall_pdf():
    # Wait a little to ensure the window is in focus
    # time.sleep(2)

    logger.info(pyautogui.position())

    # pyautogui.moveTo(548, 15)
    pyautogui.moveTo(513, 12)
    # time.sleep(1)  # Small pause
    pyautogui.click()
    # return  pyautogui.position()

# close_recall_pdf()

def check_file_exist_and_empty(file_path):
    try:
        if not os.path.exists(file_path):
            print("DMS report does not exist.")
            return False
        else:
            df = pd.read_excel(file_path, skiprows=7)
            if not df.empty and df["Customer Name"].apply(lambda x: str(x).isalpha).all():
                print("DMS report exists and has data.")
                return True
            else:
                print("There is no Jobcard details found in the DMS report.")
                return False                
    except Exception as e:
        print(f"Error while reading DMS report: {e}")
        return True
    
def check_for_empty_data_extracted(extracted_data):
    
    empty_keys = ""
    excluded_keys = {"MCP","Technical Campaign","Phone","Sub Service Type","Extended Warranty","Vehicle Sales Date"}

    # Check for empty values and append the key to the string
    for key, value in extracted_data.items():
        if key not in excluded_keys and (value is None or value == ""):
            empty_keys += key + ","

    # Trim any trailing space
    empty_keys = empty_keys.strip()
    empty_keys = empty_keys.rstrip(',')
    if empty_keys:
        empty_keys = "Empty values extracted in the fields "+empty_keys
    print("Empty keys:", empty_keys)
    return  empty_keys

def check_consolidated_report_exist_and_empty(file_path):
    try:
        if not os.path.exists(file_path):
            print("DMS report does not exist.")
            return False
        else:
            df = pd.read_excel(file_path, skiprows=1)
            if not df.empty and df["Customer Name"].apply(lambda x: str(x).isalpha).all():
                print("Consolidated report exists and has data.")
                return True
            else:
                print("There is no Jobcard details found in the Consolidated report.")
                return False                
    except Exception as e:
        print(f"Error while reading Consolidated report: {e}")
        return True

def move_screenshot(screenshot_path, log_folder):
    # Check if the log folder exists, if not, create it
    if not os.path.exists(log_folder):
        os.makedirs(log_folder)
    
    # Move the screenshot to the log folder
    destination = os.path.join(log_folder, os.path.basename(screenshot_path))
    shutil.move(screenshot_path, destination)
    print(f"Moved {screenshot_path} to {destination}")


 
# Replace with your actual OpenAI API key

 
# def extract_image_data(image_path):

#     OPENAI_API_KEY = "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"
#     client = openai.OpenAI(api_key=OPENAI_API_KEY)

#     image_prompt = """
#     Assume that you are an expert in data extraction domain and you are responsible for processing data from images...
#     """

#     image_prompt = """
#     Extract service_advisor_code, technician_code, service_advisor_name, vehicle_id, vehicle_model_id, chasis_no, vehicle_variant_id, color_id, is_extended_warranty, is_mcp, technical_campaing from the image and return the data in json format.
#     """


    
#     with open(image_path, "rb") as image_file:
#         base64_image = base64.b64encode(image_file.read()).decode("utf-8")

#     response = client.chat.completions.create(
#         model="gpt-4o-mini",
#         messages=[
#             {"role": "system", "content": "You are a helpful assistant that extracts text from images."},
#             {"role": "user", "content": [
#                 {"type": "text", "text": image_prompt},
#                 {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}}
#             ]}
#         ],
#         max_tokens=1000
#     )

#     extracted_text = response.choices[0].message.content
#     print(extracted_text)
    
#     if extracted_text.startswith("```json"):
#         extracted_text = extracted_text[7:-3].strip()

#     try:
#         json_data = json.loads(extracted_text)
#         return json_data
#     except json.JSONDecodeError:
#         return {"error": "GPT-4o-mini returned non-JSON response", "raw_response": extracted_text}
 
 



# Get the path to the current script (ab.py)

# json_path = Path(__file__).resolve().parent.parent / "Config" / "hwr 5.json"
# json_path = Path(r"C:\JobcardOpeningIntegrated") / "Config" / "hwr 5.json"
json_path = Path(r"C:\JobcardOpeningIntegrated") / "Config" / "phrasal-edition-455411-i2-d168046eb05d.json"
# json_path = Path(r"C:\JobcardOpeningIntegrated") / "Config" / "hwr 5_popular.json"

# GOOGLE_APPLICATION_CREDENTIALS - this should be json file path containing the credentials
# os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = r'E:\JobcardOpeningIntegrated\Config\hwr 5.json'  
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = str(json_path)

def extract_image_data(file_path, max_retries=3):
 
    # project and location - available in the json file containing credentials
    safety_config = [
            SafetySetting(
                category=HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT,
                threshold=HarmBlockThreshold.BLOCK_NONE,
            ),
            SafetySetting(
                category=HarmCategory.HARM_CATEGORY_HARASSMENT,
                threshold=HarmBlockThreshold.BLOCK_NONE,
            ),
            SafetySetting(
                category=HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT,
                threshold=HarmBlockThreshold.BLOCK_NONE,
            ),
            SafetySetting(
                category=HarmCategory.HARM_CATEGORY_HATE_SPEECH,
                threshold=HarmBlockThreshold.BLOCK_NONE,
            ),
        ]
   
    # vertexai.init(project='phrasal-edition-455411-i2', location='asia-south1')
    # vertexai.init(project='hwr-project-418605', location='asia-south1')
    # vertexai.init(project='zinc-ellipse-418415', location='us-central1')
    vertexai.init(project='phrasal-edition-455411-i2', location='us-central1')
    # model = GenerativeModel(model_name="gemini-1.5-flash-002",safety_settings=safety_config)
    # model = GenerativeModel(model_name="gemini-2.5-flash",safety_settings=safety_config)
    model = GenerativeModel(model_name="gemini-2.0-flash-lite",safety_settings=safety_config)
    with open(file_path, "rb") as image_file:
        image_data = image_file.read()
 
    prompt_text = """
    Assume that you are an expert in data extraction domain and you are responsible for processing data from images...
    """
 
    prompt_text = """
    Extract service_advisor_code, service_advisor_name, vehicle_model_id, vehicle_variant_id, color_id, omr, sub_service_type, service_type_id, is_extended_warranty, is_mcp, technical_campaing from the image and return the data in json format.
    """ 
 
    # prompt_text = """Extract tabular data in json format"""
 
    image_prompt = [Part.from_data(image_data, mime_type="image/jpeg"),prompt_text]
   
    dict_response = None  # Initialize dict_response here
 
    try:
        response = model.generate_content(image_prompt)

        if response.usage_metadata:
            print("\n🔢 Token Usage:")
            print("Prompt tokens:", response.usage_metadata.prompt_token_count)
            print("Response tokens:", response.usage_metadata.candidates_token_count)
            print("Total tokens:", response.usage_metadata.total_token_count)
        else:
            print("No usage metadata available.")

        data = response.text
        data = data.replace("```json", '').replace("```", '')
        # python_obj_data = ast.literal_eval(data)
        # cleaned_data = json.dumps(python_obj_data, indent=2)
        dict_response = json.loads(data)
    except (ValueError, SyntaxError,json.JSONDecodeError) as e:
        return {"error": "GPT-4o-mini returned non-JSON response", "raw_response": data}
 
    if dict_response is None:
        raise RuntimeError("Failed to obtain a valid response.")
 
    return dict_response
 
# file_path = r"C:\Users\Akhil Nandha Nandha\Downloads\Tabular-data-used-in-examples.png"
# output = extract_image_data(file_path)
# print(output)
 


    # return dict_response
 
# file_path = r"E:\JobcardOpeningIntegrated\DMS_captured\jobcard_image_JC24021092.png"
# output =
# extract_image_data(file_path)
# print(output)
# print(type(output))

# image_path = r"E:\JobcardOpeningIntegrated\DMS_captured\jobcard_image_JC24021092.png"
# output = extract_image_data(image_path)
# print(json.dumps(output, indent=4))

def remove_all_csv_files_from_downloads():
    # Define the Downloads folder path
    downloads_folder = os.path.expanduser("~/Downloads")  # Works on Windows, macOS, and Linux

    # Get a list of all .csv files in the Downloads folder
    csv_files = glob.glob(os.path.join(downloads_folder, "*.csv"))

    # Delete each CSV file
    for file in csv_files:
        try:
            os.remove(file)
            print(f"Deleted: {file}")
        except Exception as e:
            print(f"Error deleting {file}: {e}")

    print("Cleanup complete.")

def read_dms_location_from_config(file_path):
    try:
        
        df = pd.read_excel(file_path, sheet_name='Sheet1')
        location = df.iloc[0, 4]  
        config_time_out = df.iloc[0, 16] 
        
        try:
            config_time_out = float(config_time_out)
        except (ValueError, TypeError):
            config_time_out = 45

        return location, config_time_out
    except Exception as e:
        print(f"An error occurred: {e}")
        return None, None

def clear_log_folder(log_folder):
    try:
        resolved_path = os.path.abspath(os.path.normpath(log_folder))
        print(f"Resolved log folder path: {resolved_path}")
        
        for filename in os.listdir(log_folder):
            file_path = os.path.join(log_folder, filename)
            if os.path.isfile(file_path):
                os.remove(file_path)
                print(f"Deleted: {file_path}")
        print("All files deleted successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")
# log_folder = r"E:\JobcardOpeningIntegrated\Log"
# clear_log_folder(log_folder)

def check_for_the_dms_report_present_in_the_download_folder():
    downloads_path = str(Path.home() / "Downloads")
    today = datetime.date.today()

    max_attempts = 5
    attempt = 0

    while attempt < max_attempts:
        for filename in os.listdir(downloads_path):
            if filename.endswith(".csv"):
                file_path = os.path.join(downloads_path, filename)
                modified_date = datetime.date.fromtimestamp(os.path.getmtime(file_path))
                if modified_date == today:
                    print(f"Found today's CSV: {filename}")
                    return True
        attempt += 1
        print(f"Attempt {attempt} - CSV not found. Retrying in 2 seconds...")
        time.sleep(3)

    print("No CSV file downloaded today was found after 5 attempts.")
    return False

