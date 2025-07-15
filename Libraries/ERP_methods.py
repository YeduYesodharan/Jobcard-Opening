import re
from datetime import datetime
import pandas as pd
from pathlib import Path
from robot.api  import logger

def extract_name_from_string(input_string):
    """
    Extracts the name from the given input string.

    Args:
    input_string (str): The input string containing the name information.

    Returns:
    str: The extracted name, if found. Otherwise, returns None.
    """
    pattern = r"name='([^']+)'"
    match = re.search(pattern, input_string)
    if match:
        return match.group(1)
    else:
        return None
    


def get_title_starting_with(window_info, prefix):
    """
    Function to get the title starting with the specified prefix from the window information dictionary.
    
    Args:
    - window_info (dict): Information dictionary of the window containing 'title' key.
    - prefix (str): Prefix string to match with the beginning of the title.

    Returns:
    - str or None: The title starting with the specified prefix, or None if not found.
    """
    title = window_info.get('title', '')
    if title.startswith(prefix):
        return title
    return None



# import re
# from datetime import datetime

# def extract_date_and_time(expected_delivery_date):
#     # If the input is a datetime object, convert it to a string first
#     if isinstance(expected_delivery_date, datetime):
#         expected_delivery_date = expected_delivery_date.strftime("%d-%m-%Y %H:%M:%S")
    
#     # Normalize spacing by removing extra spaces between date and time
#     expected_delivery_date = ' '.join(expected_delivery_date.split())

#     # Use regex to separate date and time part (handles multiple spaces and no space)
#     match = re.match(r"(\d{2}-\d{2}-\d{4})\s*(\d{2}:\d{2}(:\d{2})?)", expected_delivery_date)

#     if not match:
#         raise ValueError(f"Date format '{expected_delivery_date}' does not match expected formats.")

#     # Extract date and time parts
#     date_part = match.group(1)  # 'DD-MM-YYYY'
#     time_part = match.group(2)  # 'HH:MM' or 'HH:MM:SS'

#     # Convert to datetime object for further formatting if needed
#     if len(time_part) == 5:  # Format like 'HH:MM'
#         delivery_datetime = datetime.strptime(date_part + " " + time_part, "%d-%m-%Y %H:%M")
#     else:  # Format like 'HH:MM:SS'
#         delivery_datetime = datetime.strptime(date_part + " " + time_part, "%d-%m-%Y %H:%M:%S")
    
#     # Extract the date and time in the required format
#     promised_date = delivery_datetime.strftime("%d/%m/%Y")  # Date part
#     promised_time = delivery_datetime.strftime("%H:%M:%S")  # Time part (always in HH:MM:SS format)

#     # Return the desired format: "variable1: {date} v2: {time}"
#     return promised_date, promised_time
from datetime import datetime
import re

def extract_date_and_time(expected_delivery_date):
    # If the input is a datetime object, convert it to a string first
    if isinstance(expected_delivery_date, datetime):
        expected_delivery_date = expected_delivery_date.strftime("%d-%m-%Y %H:%M:%S")
    
    # Normalize spacing by removing extra spaces between date and time
    expected_delivery_date = ' '.join(expected_delivery_date.split())

    # Use regex to separate date and time part (handles multiple spaces and no space)
    match = re.match(r"(\d{2}-\d{2}-\d{4})\s*(\d{2}:\d{2}(:\d{2})?)", expected_delivery_date)

    if not match:
        raise ValueError(f"Date format '{expected_delivery_date}' does not match expected formats.")

    # Extract date and time parts
    date_part = match.group(1)  # 'DD-MM-YYYY'
    time_part = match.group(2)  # 'HH:MM' or 'HH:MM:SS'

    # Convert to datetime object for further formatting if needed
    if len(time_part) == 5:  # Format like 'HH:MM'
        delivery_datetime = datetime.strptime(date_part + " " + time_part, "%d-%m-%Y %H:%M")
    else:  # Format like 'HH:MM:SS'
        delivery_datetime = datetime.strptime(date_part + " " + time_part, "%d-%m-%Y %H:%M:%S")
    
    # Extract the date and time in the required format
    promised_date = delivery_datetime.strftime("%d-%m-%Y")  # Date part in DD-MM-YYYY format
    promised_time = delivery_datetime.strftime("%H:%M:%S")  # Time part in HH:MM:SS format

    # Return the desired format: "variable1: {date} v2: {time}"
    return promised_date, promised_time




# def validate_service_model_code(excel_path, service_model_code_value):
   
    
#     # Load the Excel file into a DataFrame
#     df = pd.read_excel(excel_path)

#     # Check if the necessary columns exist in the DataFrame
#     if not all(col in df.columns for col in ['Serial No', 'Code', 'ERP SERVICE MODEL CODE', 'DMS Vehicle Model']):
#         raise ValueError("Excel file must contain the columns: 'Serial No', 'Code', 'ERP SERVICE MODEL CODE', 'DMS Vehicle Model'.")

#     # Iterate through each row in the DataFrame
#     for index, row in df.iterrows():
#         # If the 'DMS Vehicle Model' column matches the service model code value
#         if row['DMS Vehicle Model'] == service_model_code_value:
#             # Return the corresponding 'Code' value for that row
#             return row['ERP SERVICE MODEL CODE']
    
#     # Return None if no match was found
#     return None

def validate_service_model_code(excel_path, service_model_code_value):
   
    
    # Load the Excel file into a DataFrame
    df = pd.read_excel(excel_path)

    # Check if the necessary columns exist in the DataFrame
    if not all(col in df.columns for col in ['Serial No', 'Code', 'ERP SERVICE MODEL CODE', 'DMS Vehicle Model']):
        raise ValueError("Excel file must contain the columns: 'Serial No', 'Code', 'ERP SERVICE MODEL CODE', 'DMS Vehicle Model'.")

    # Iterate through each row in the DataFrame
    for index, row in df.iterrows():
        # If the 'DMS Vehicle Model' column matches the service model code value
        if row['DMS Vehicle Model'] == service_model_code_value:
            # Return the corresponding 'Code' value for that row
            return row['ERP SERVICE MODEL CODE']
    
    # Return None if no match was found
    return 'no match found for vehicle service model'



import pandas as pd
import math

def sort_and_get_service_type_from_excel(excel_path, type_of_service, service_code):
    # Step 1: Load the Excel file using pandas
    df = pd.read_excel(excel_path)

    # Step 2: Clean column names (strip leading/trailing spaces)
    df.columns = df.columns.str.strip()

    # Step 3: Ensure type_of_service is a string and convert it to uppercase
    if not isinstance(type_of_service, str):
        raise ValueError(f"Expected 'type_of_service' to be a string, but got {type(type_of_service)}")
    
    type_of_service = type_of_service.upper()

    # Step 4: Check if required columns exist in the DataFrame
    required_columns = ['Type Of Service', 'Service Type', 'DMS Service Description', 'DMS Service Code']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")

    # Step 5: Filter the rows based on the Type Of Service column
    filtered_df = df[df['Type Of Service'] == type_of_service]

    # Step 6: Sort the filtered data by the 'DMS Service Code' column
    sorted_df = filtered_df.sort_values(by='DMS Service Code')

    # Step 7: Find the corresponding Service Type for the given Service Code
    if service_code is None:
        return 'no match found for service type'
    else:
        service_code = service_code.replace(" ", "")        
        result = sorted_df[sorted_df['DMS Service Code'].str.replace(" ", "", regex=True) == service_code]
        if not result.empty: 

            logger.info(result.iloc[0]['Service Type'])
            nan_check = False
            if isinstance(result.iloc[0]['Service Type'] , float):
                nan_check = math.isnan(result.iloc[0]['Service Type'])

            if not nan_check: 
                if not result.empty:        
                # Step 8: If a match is found, return the Service Type, otherwise return None
                # if not result.empty:
                    return result.iloc[0]['Service Type']  # Access the 'Service Type' column from the first row
                else:
                    return 'no match found for service type'
            else:
                return 'no match found for service type'
        else:
            return 'no match found for service type'

# import openpyxl

# def Advisor_name_from_dms(excel_file_path, service_advisor):
#     # Load the Excel workbook
#     wb = openpyxl.load_workbook(excel_file_path)
    
#     # Assuming data is on the first sheet
#     sheet = wb.active
    
#     # Iterate over the rows of the sheet
#     for row in sheet.iter_rows(min_row=2, values_only=True):  # start from row 2 assuming row 1 has headers
#         dms_name = row[4]  # DMS Name column (5th column, index 4)
#         name = row[1]  # Name column (2nd column, index 1)

#         # Check if the DMS Name matches the input variable
#         if dms_name == service_advisor:
#             return name
    
#     return None  # If no match found, return None

import openpyxl

def Advisor_name_from_dms(excel_file_path, service_advisor):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(excel_file_path)
    
    # Assuming data is on the first sheet
    sheet = wb.active
    
    # Iterate over the rows of the sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):  # start from row 2 assuming row 1 has headers
        dms_name = row[4]  # DMS Name column (5th column, index 4)
        name = row[1]  # Name column (2nd column, index 1)

        # Check if the DMS Name matches the input variable
        if dms_name == service_advisor:
            return name
    
    return 'no match found for service advisor'  # If no match found, return None

from datetime import datetime

def convert_date_format(date_input):
    # If the input is already a datetime object, directly format it
    if isinstance(date_input, datetime):
        return date_input.strftime("%d/%m/%Y")
    
    # If the input is a string, convert it to a datetime object first
    date_obj = datetime.strptime(date_input, "%d-%m-%Y")
    return date_obj.strftime("%d/%m/%Y")


# def increase_cordinates(region, x, y, w, h):
#     x = int(x.strip('"'))
#     y = int(y.strip('"'))
#     w = int(w.strip('"'))
#     h = int(h.strip('"'))
   
#     region[0] += x
#     region[1] += y
#     region[2] += w
#     region[3] += h
#     return region


import psutil




import os
import glob

def open_application():
    app_name = "Wings 23E Launcher"
    search_dirs = [
        os.path.join(os.environ.get("USERPROFILE", ""), "Desktop"),
        os.path.join(os.environ.get("PUBLIC", ""), "Desktop"),
        os.path.join(os.environ.get("ProgramFiles", "")),
        os.path.join(os.environ.get("ProgramFiles(x86)", "")),
        # Add AppData and Start Menu
        os.path.join(os.environ.get("APPDATA", ""), "Microsoft\\Windows\\Start Menu"),
        os.path.join(os.environ.get("LOCALAPPDATA", "")),
        os.path.join(os.environ.get("ProgramData", ""), "Microsoft\\Windows\\Start Menu"),
        # Add additional drives if needed
        "D:\\", "E:\\", "F:\\"  # Example: You can add other drives here
    ]
    
    # Search for the .lnk or .exe file
    for search_dir in search_dirs:
        app_path = glob.glob(search_dir + f"\\*{app_name}*.lnk")
        if app_path:
            os.startfile(app_path[0])
            return

        app_path = glob.glob(search_dir + f"\\*{app_name}*.exe")
        if app_path:
            os.startfile(app_path[0])
            return

    print(f"Application '{app_name}' not found.")






# import psutil

# def close_application():
#     app_name_part = "Wings"  # Part of the name to search for (modify if needed)
    
#     # Flag to check if we found the application
#     found = False
    
#     # Iterate through all running processes
#     for proc in psutil.process_iter(['pid', 'name']):
#         # Check if the application part of the name is in the process name (case-insensitive)
#         if app_name_part.lower() in proc.info['name'].lower():
#             found = True
#             print(f"Found process with name containing '{app_name_part}', PID: {proc.info['pid']}")
#             try:
#                 proc.kill()  # Forcefully terminate the process
#                 print(f"Application with PID {proc.info['pid']} has been closed.")
#             except (psutil.NoSuchProcess, psutil.AccessDenied):
#                 print(f"Couldn't terminate process with PID {proc.info['pid']}.")

#     if not found:
#         print(f"No process found containing '{app_name_part}'.")

# # Run the function
# close_application()

import psutil

def close_application():
    app_name_part = "Wings"  # Part of the name to search for (modify if needed)
    
    # Flag to check if we found the application
    found = False
    
    # Iterate through all running processes
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            # Check if the application part of the name is in the process name (case-insensitive)
            if app_name_part.lower() in proc.info['name'].lower():
                found = True
                print(f"Found process with name containing '{app_name_part}', PID: {proc.info['pid']}")
                
                # Additional check for the specific launcher process if known
                if 'Launcher' in proc.info['name']:
                    print(f"Found launcher process with name {proc.info['name']}, PID: {proc.info['pid']}")
                    
                proc.kill()  # Forcefully terminate the process
                print(f"Application with PID {proc.info['pid']} has been closed.")
        
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess) as e:
            print(f"Couldn't terminate process with PID {proc.info['pid']}: {e}")

    if not found:
        print(f"No process found containing '{app_name_part}'.")

# Run the function
# close_application()









import tkinter as tk
from tkinter import messagebox

def show_message_box(title, message):
    # Create a root window, but do not display it
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    # Force the message box to appear in the foreground by calling 'lift' on the root
    root.lift()  # Bring the root window to the front
    root.attributes("-topmost", True)  # Ensure it stays on top

    # Show the message box
    messagebox.showinfo(title, message)

    # After the message box is closed, destroy the root window
    root.destroy()






import os

def check_file_exists(file_path):
    """
    Checks if a file exists at the specified path.

    :param file_path: The path to the file to check.
    :return: True if the file exists, False otherwise.
    """
    return os.path.isfile(file_path)


import pandas as pd
import re
import os

def extract_and_correct_registration(excel_path):
    # Check if the provided path is a valid file
    if not isinstance(excel_path, (str, bytes, os.PathLike)):
        return "The provided path is not a valid string, bytes, or os.PathLike object."

    # Check if the file exists at the provided path
    if not os.path.isfile(excel_path):
        return f"The file at '{excel_path}' does not exist."
    
    try:
        # Load the Excel file into a pandas DataFrame
        df = pd.read_excel(excel_path)
    except Exception as e:
        return f"Error loading the Excel file: {str(e)}"
    
    # Check if the 'Registration no' column exists
    if "registration no" not in df.columns:
        return "The column 'Registration no' was not found."
    
    # Function to correct the registration number
    def correct_registration_number(reg_number):
        reg_number = reg_number.strip().upper()  # Convert to upper case for case insensitivity
        
        # Regex to match a state code (2 uppercase letters) followed by one or two digits and alphanumeric characters
        match = re.match(r"^([A-Z]{2})(\d{1,2})([A-Z0-9]+)$", reg_number)
        
        if match:
            state_code = match.group(1)
            state_digit = match.group(2)
            vehicle_number = match.group(3)
            
            # Add leading zero if the state code has only one digit
            if len(state_digit) == 1:
                state_digit = '0' + state_digit
            
            # Return the corrected registration number
            return f"{state_code}{state_digit}{vehicle_number}"
        
        # If the format is incorrect, return the original registration number
        return reg_number
    
    # Apply the correction function to the 'Registration no' column and update it
    df['registration no'] = df['registration no'].apply(correct_registration_number)
    
    # Save the updated DataFrame to the same Excel file (overwrite the original)
    try:
        df.to_excel(excel_path, index=False)
    except Exception as e:
        return f"Error saving the updated Excel file: {str(e)}"
    
    # Return the path of the same updated Excel file
    return excel_path






# Example usage:
# excel_path = "C://Popular ERP Process - Copy//Input//consolidated_report.xlsx"    # Update with your file path
# corrected_values = extract_and_correct_registration(excel_path)
# print(corrected_values)


import pandas as pd

def extract_and_combine_addresses(excel_path):
    # Load the Excel file
    df = pd.read_excel(excel_path)

    # Check if the required columns exist
    required_columns = ['Address 1', 'Address 2', 'Address 3']
    if not all(col in df.columns for col in required_columns):
        return "The required columns 'Address 1', 'Address 2', or 'Address 3' were not found."

    # Combine the values from 'Address 1', 'Address 2', and 'Address 3' into one column with spaces
    df['Combined Address'] = df['Address 1'].fillna('') + ' ' + df['Address 2'].fillna('') + ' ' + df['Address 3'].fillna('')
    
    # Combine all the addresses into a single string, separated by a newline
    combined_addresses_text = '\n'.join(df['Combined Address'].tolist())
    
    # Return the combined addresses as a single text value
    return combined_addresses_text


# # Example usage:
# excel_path = "path_to_your_excel_file.xlsx"  # Update with your file path
# combined_addresses = extract_and_combine_addresses(excel_path)
# print(combined_addresses)

def extract_demand_codes(demand_codes_str):
    # Check if the input is None
    if demand_codes_str is None:
        return []  # Or handle it in any way that fits your needs

    # Split the string by commas and strip any extra spaces
    demand_codes_list = [code.strip() for code in demand_codes_str.split(',')]
    return demand_codes_list


import pandas as pd

def map_type_of_service(mapping_excel_path,service_type_code,sub_service_type):
    # Load the input sheet containing 'Service Type Code' or 'Sub Service Type'
    #df = pd.read_excel(excel_path)
    #test
    type_of_service = None
    # Load the mapping sheet containing 'DMS Service Code' and 'Type Of Service'

    mapping_df = pd.read_excel(mapping_excel_path)
    
    # Check if the necessary columns exist in both dataframes
    #if "Service Type Code" not in df.columns and "Sub Service Type" not in df.columns:
     #   return "The column 'Service Type Code' or 'Sub Service Type' was not found in the input sheet."
    
    if "DMS Service Code" not in mapping_df.columns or "Type Of Service" not in mapping_df.columns:
        return "The columns 'DMS Service Code' or 'Type Of Service' were not found in the mapping sheet."
    
    # First check using 'Service Type Code'
    #for service_type_code in df.get('Service Type Code', []):
        # Try to find the corresponding 'DMS Service Code' in the mapping sheet
    service_type_code = service_type_code.replace(" ", "")  
    mapping_row = mapping_df[mapping_df['DMS Service Code'].str.replace(" ", "", regex=True) == service_type_code]
        
    if not mapping_row.empty:
            # If a match is found, get the corresponding 'Type Of Service' value
        type_of_service = mapping_row['Type Of Service'].iloc[0]
        return type_of_service
    
    # If no match found with 'Service Type Code', check using 'Sub Service Type'
    #if "Sub Service Type" in df.columns:
    if type_of_service is None:
        #for sub_service_type in df['Sub Service Type']:
            # Try to find the corresponding 'DMS Service Code' in the mapping sheet
            sub_service_type = sub_service_type.replace(" ", "")  
            mapping_row = mapping_df[mapping_df['DMS Service Code'].str.replace(" ", "", regex=True) == sub_service_type]
            
            if not mapping_row.empty:
                # If a match is found, get the corresponding 'Type Of Service' value
                type_of_service = mapping_row['Type Of Service'].iloc[0]
                return type_of_service

    # If no match is found with either column, return a message indicating so
 
    return "no match found for type of service"



import os
from datetime import datetime

def append_current_datetime_to_path(base_path):
    # Get the current date and time
    current_datetime = datetime.now()
    # Format the date and time in the desired format
    formatted_datetime = current_datetime.strftime("%d-%m-%Y_%I-%M%p")
    
    # Split the base path into directory and file name
    dir_name, base_file = os.path.split(base_path)
    
    # Create the new path by appending the formatted date and time
    new_file_name = f"{base_file}_{formatted_datetime}"
    new_path = os.path.join(dir_name, new_file_name)
    
    return new_path

# Example usage
# input_path = "E:\JobcardOpeningIntegrated\Recall Reports\Report"
# new_path = append_current_datetime_to_path(input_path)
# print("New path with date and time:", new_path)


import pandas as pd

def Recall_Tagged_or_not(excel_path, branch, circular_number, chassis_number, registration_number):
    # Clean the input values by ensuring they are not None before stripping leading and trailing spaces
    branch = (branch or "").strip()
    circular_number = (circular_number or "").strip()
    chassis_number = (chassis_number or "").strip()
    registration_number = (registration_number or "").strip()

    # Load the Excel file, specifying that the header starts from the 4th row (index 3)
    df = pd.read_excel(excel_path, header=3)

    # Clean column names by stripping extra spaces
    df.columns = df.columns.str.strip()

    # Clean the values in relevant columns by stripping extra spaces, handling None
    df['Branch'] = df['Branch'].fillna("").str.strip()
    df['Circular Number'] = df['Circular Number'].fillna("").str.strip()
    df['Chassis Number'] = df['Chassis Number'].fillna("").str.strip()
    df['Registration Number'] = df['Registration Number'].fillna("").str.strip()

    # Check if the row exists with all the given conditions (Branch, Circular Number, Chassis Number, and Registration Number)
    filtered_rows = df[
        (df['Branch'] == branch) &
        (df['Circular Number'] == circular_number) &
        (df['Chassis Number'] == chassis_number) &
        (df['Registration Number'] == registration_number)
    ]

    # If the row is found based on the complete conditions, return "Recall tagged"
    if not filtered_rows.empty:
        return "recall tagged"
    
    # If row not found, iterate again with just Branch and Circular Number
    filtered_rows_by_branch_circular = df[
        (df['Branch'] == branch) &
        (df['Circular Number'] == circular_number)
    ]
    
# If no rows are found with just Branch and Circular Number, return the appropriate message
    if filtered_rows_by_branch_circular.empty:
        return "circular number or branch doesn't exist in recall report"

    # Check if we have multiple rows with the same Voucher No.
    voucher_numbers = filtered_rows_by_branch_circular['Voucher No'].unique()

    # Return the Voucher No. if there's only one unique Voucher No
    if len(voucher_numbers) == 1:
        return voucher_numbers[0]



# # Example usage
# excel_path = "E:\\JobcardOpeningIntegrated\\Recall Reports\\Report_10-03-2025_09-20PM.xlsx"  # Replace with your actual Excel file path
# branch = "SLM_MGM"
# circular_number = "D-59/2023"
# chassis_number = "MA3RYHL1SNG113044"
# registration_number = "KL56X4399"

# result = Recall_Tagged_or_not(excel_path, branch, circular_number, chassis_number, registration_number)
# print(result)  


def check_file_exist_and_empty(file_paths):
    try:
        if not file_paths:  # Check if the list is empty
            print("No files found in the specified folder.")
            return False
        else:
            print(f"{len(file_paths)} file(s) found in the folder.")
            return True
    except Exception as e:
        print(f"Error while checking files: {e}")
        return False

    
def recall_process_related_folder_creation(curr_date):
    try:
        #----------------- Getting the root folder path ----------------
        project_root = get_process_root_directory()
        #---------------------------------------------------------------

        #-----Creating RecallReports folder inside root folder -----
        # Define the new folder path inside the root folder
        recall_reports_folder = project_root / "RecallReports"

        # Create the folder if it doesn't exist
        recall_reports_folder.mkdir(parents=True, exist_ok=True)
        print(f"Project Root Directory: {recall_reports_folder}")

        #--------------- Creating current date folder inside "RecallReports"--------
        # Define the new folder path with today's date
        recall_date_folder = recall_reports_folder / curr_date

         # Create the folders if they don't exist
        recall_date_folder.mkdir(parents=True, exist_ok=True)

        print(f"Folder created (or already exists) at: {recall_date_folder}")
        #---------------------------------------------------------------------

        #-----Creating process_related_folders folder inside root folder -----
        # Define the new folder path inside the root folder
        process_related_folder = project_root / "ProcessRelatedFolders"
 
        # Create the folder if it doesn't exist
        process_related_folder.mkdir(parents=True, exist_ok=True)
        print(f"Project Root Directory: {process_related_folder}")
        #---------------------------------------------------------------------
 
        #--------------- Creating current date folder inside "process_related_folders"--------
        # Define the new folder path with today's date
        date_folder = process_related_folder / curr_date
 
        # Create the folders if they don't exist
        date_folder.mkdir(parents=True, exist_ok=True)
 
        print(f"Folder created (or already exists) at: {date_folder}")
        #------------------------------------------------------------------------

        # -------  "ProcessedDMS"inside current date folder -----
        # Define the subfolders to create inside the date folder
        subfolders = ["ProcessedDMS"]
 
        # Create the folders if they don't exist
        for subfolder in subfolders:
            (date_folder / subfolder).mkdir(parents=True, exist_ok=True)
 
        print(f"Folders created inside: {date_folder}")
        #--------------------------------------------------------
    
        # ----- Creating "Completed" inside "ProcessedDMS" -----
        completed_folder = date_folder / "ProcessedDMS" / "Completed"

        # Create the "Completed" folder if it doesn't exist
        completed_folder.mkdir(parents=True, exist_ok=True)
        print(f"Completed folder created at: {completed_folder}")
        #--------------------------------------------------------

        return completed_folder
    except Exception as e:
        print(f"Error while creating the folders: {str(e)}")
        raise Exception(e)

# def get_process_root_directory():
#     #----------------- Getting the root folder path ----------------
#     # Get the absolute path of the current script
#     current_file_path = Path(__file__).resolve()

#     # Move up one level to reach "JobcardOpeningProcess"
#     project_root = current_file_path.parents[1]

#     print(f"Project Root Directory: {project_root}")
#     #---------------------------------------------------------------
#     return project_root

def get_process_root_directory():
    # Return a fixed path for the project root
    project_root = Path(r"C:\JobcardOpeningIntegrated")
    print(f"Project Root Directory: {project_root}")
    return project_root

import os
import shutil
import glob
import pandas as pd
from datetime import datetime

def merge_and_move_files(source_path, destination_path):
    # Check if the source path exists
    if not os.path.exists(source_path):
        raise Exception(f"source path does not exist.")
    
    # Find all .xlsx files in the source directory
    xlsx_files = glob.glob(os.path.join(source_path, '*.xlsx'))
    
    if not xlsx_files:
        raise Exception(f"No files with Recalls found in folder")
    
    # Check if the destination path exists, create if not
    if not os.path.exists(destination_path):
        os.makedirs(destination_path)

    # Move the files to the destination folder
    for file in xlsx_files:
        shutil.move(file, os.path.join(destination_path, os.path.basename(file)))
    
    # List the moved files in the destination folder
    moved_files = glob.glob(os.path.join(destination_path, '*.xlsx'))
    
    # Read and merge the xlsx files into a single DataFrame
    merged_df = pd.concat([pd.read_excel(file) for file in moved_files], ignore_index=True)
    
    # Generate the filename for the consolidated report with current date and time
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    consolidated_filename = f"Consolidated_Jobcard_Report_Recalls_{current_time}.xlsx"
    
    # Save the merged DataFrame as a new Excel file
    consolidated_filepath = os.path.join(destination_path, consolidated_filename)
    merged_df.to_excel(consolidated_filepath, index=False)
    
    # Delete the individual moved files after merging
    for file in moved_files:
        os.remove(file)
    
    # Return the path of the consolidated file
    # print(consolidated_filepath)
    return consolidated_filepath

def prepare_file_name():
    current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    recall_report_filename = f"Report_{current_time}.xlsx"
    return recall_report_filename


import pandas as pd

def get_erp_branch_location_code(excel_path, search_value):
    try:
        # Load the Excel file into a DataFrame
        df = pd.read_excel(excel_path)
        
        # Check if the columns exist in the DataFrame
        if 'DMS Location description' in df.columns and 'ERP location code' in df.columns:
            # Search for the value in the 'DMS Location description' column
            match_row = df[df['DMS Location description'] == search_value]
            
            # If a match is found, return the corresponding 'ERP location code'
            if not match_row.empty:
                return match_row['ERP location code'].iloc[0]
            else:
                return "branch code not available"
        else:
            return "Required columns not found in the Excel file"

    except Exception as e:
        return f"An error occurred: {e}"


def contains_substring(main_string, substring):
    return substring in main_string  

def copy_report_to_destination_folder(merged_report, destination_folder):
    # Ensure destination folder exists
    os.makedirs(destination_folder, exist_ok=True)
 
    consolidated_report_file_to_check = os.path.join(destination_folder, "consolidated_jobcard_report.xlsx")
 
    # Check if the file exists in the destination folder
    if not os.path.exists(consolidated_report_file_to_check):
        shutil.copy(merged_report, consolidated_report_file_to_check)
        print(f" {merged_report} File copied to {consolidated_report_file_to_check}")
    else:
        print(f"{merged_report} File already exists: {consolidated_report_file_to_check}")
        # Need to merge new rows from the source consolidated report with timestamp to existing consolidated report without stamp in the Results/current_date folder
 
        # Read existing file
        df_existing = pd.read_excel(consolidated_report_file_to_check)
        df_new = pd.read_excel(merged_report)
 
        # Append new data to existing data
        df_combined = pd.concat([df_existing, df_new], ignore_index=True)
 
        # Save back to consolidated file
        df_combined.to_excel(consolidated_report_file_to_check, index=False)
        print("New JC rows appended to existing consolidated_jobcard_report.xlsx")

 
    return  consolidated_report_file_to_check

import os
import shutil

def Move_to_Completed_Folder(source_path,destination_path):
    try:
        # Check if the file has a '.xlsx' extension
        if source_path.endswith('.xlsx'):
            # Get the directory and base file name (without extension)
            # directory, filename = os.path.split(source_path)
            # filename_without_extension = os.path.splitext(filename)[0]

            # Append 'Completed' to the file name
            # new_filename = filename_without_extension + 'Completed'
            # new_path = os.path.join(directory, new_filename)

            # Move the file to the new path
            shutil.move(source_path, destination_path)
            # print(f"File moved successfully from {source_path} to {new_path}")
        else:
            print("The provided file is not an '.xlsx' file.")
    
    except Exception as e:
        print(f"Error occurred while moving the file: {e}")


def clean_string(input_str):
    # Remove all spaces and convert to lowercase
    cleaned = input_str.replace(" ", "").lower()
    return cleaned

# import re
 
# def correct_registration_number_from_var(reg_number):
#     """
#     Accepts a single registration number string and returns the corrected version.
#     """
#     if not isinstance(reg_number, str):
#         return "Input must be a string."
 
#     reg_number = reg_number.strip().upper()  # Convert to uppercase and remove surrounding whitespace
 
#     # Match pattern: two letters + 1-2 digits + alphanumeric (vehicle number)
#     match = re.match(r"^([A-Z]{2})(\d{1,2})([A-Z0-9]+)$", reg_number)
   
#     if match:
#         state_code = match.group(1)
#         state_digit = match.group(2)
#         vehicle_number = match.group(3)
       
#         # Add leading zero if single digit
#         if len(state_digit) == 1:
#             state_digit = '0' + state_digit
       
#         return f"{state_code}{state_digit}{vehicle_number}"
 
#     # Return original if it doesn't match expected pattern
#     return reg_number  

import os
import signal
import subprocess
import platform

def close_browser_processes():
    browsers = ['msedge', 'chrome', 'firefox', 'iexplore']  # Add any others if needed
    current_os = platform.system()

    for browser in browsers:
        if current_os == 'Windows':
            # Use taskkill to kill processes on Windows
            subprocess.call(f'taskkill /F /IM {browser}.exe', stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, shell=True)
        elif current_os in ['Linux', 'Darwin']:
            # For Linux or macOS
            subprocess.call(f'pkill -f {browser}', shell=True)

import re

def extract_valid_mobile(text):
    match = re.search(r'\b\d{10}\b', text)
    return match.group(0) if match else ''

# import openpyxl
# from datetime import datetime

# def get_pickup_details(excel_path, jobcard_no):
#     """
#     Searches for a job card number in the given Excel file (case-insensitive) and returns the entire row as a dictionary.
#     The 'Pick Up Date' field will be formatted to 'dd-mm-yyyy' (if it is a datetime object).

#     Args:
#         excel_path (str): Path to the Excel file
#         jobcard_no (str): Job Card Number to search for (e.g., JC25002585 or jc25002585)

#     Returns:
#         dict: Row data as dictionary if found, else an empty dict
#     """
#     wb = openpyxl.load_workbook(excel_path, data_only=True)
#     sheet = wb.active

#     headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
#     jobcard_no = str(jobcard_no).strip().lower()

#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         current_jobcard = str(row[0]).strip().lower()
#         if current_jobcard == jobcard_no:
#             row_dict = dict(zip(headers, row))

#             # Format 'Pick Up Date' to 'dd-mm-yyyy' if it's a datetime
#             pickup_date = row_dict.get("Pick Up Date")
#             if isinstance(pickup_date, datetime):
#                 row_dict["Pick Up Date"] = pickup_date.strftime('%d-%m-%Y')

#             return row_dict

#     return {}

# import openpyxl
# from datetime import datetime

# def get_pickup_details(excel_path, jobcard_no):
#     """
#     Searches for a job card number in the Excel file and returns row data.
#     - If required pickup fields (except Pick Up Driver) are missing, returns a message.
#     - If Pick Up Driver starts with 'NO PICKUP' or 'EMPLOYEE PICKUP' (case-sensitive), returns a specific message.

#     Returns:
#         tuple: (row_dict, status_message)
#     """
#     wb = openpyxl.load_workbook(excel_path, data_only=True)
#     sheet = wb.active

#     headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
#     jobcard_no = str(jobcard_no).strip().lower()

#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         current_jobcard = str(row[0]).strip().lower()
#         if current_jobcard == jobcard_no:
#             row_dict = dict(zip(headers, row))

#             # Check for required fields (excluding Pick Up Driver)
#             required_fields = ["Pick Up Date", "Pick Up Time", "Distance in KM"]
#             for field in required_fields:
#                 value = row_dict.get(field)
#                 if value is None or str(value).strip() == "":
#                     return row_dict, 'Some mandatory values except Pickup Driver is missing'

#             # Check Pick Up Driver
#             pickup_driver = row_dict.get("Pick Up Driver")
#             if pickup_driver is None or str(pickup_driver).strip() == "":
#                 return row_dict, 'Pickup Driver is missing'

#             # Case-sensitive check
#             pickup_driver_value = str(pickup_driver).strip()
#             if pickup_driver_value.startswith("EMPLOYEE PICKUP"):
#                 return row_dict, 'employee pickup'
            
#             # No-Pickup check
#             pickup_driver_value = str(pickup_driver).strip()
#             if pickup_driver_value.startswith("NO PICKUP"):
#                 return row_dict, 'no pickup'

#             # Format 'Pick Up Date' if it's a datetime
#             pickup_date = row_dict.get("Pick Up Date")
#             if isinstance(pickup_date, datetime):
#                 row_dict["Pick Up Date"] = pickup_date.strftime('%d-%m-%Y')

#             return row_dict, 'fields extracted'

#     return {}, None

import openpyxl
from datetime import datetime

def get_pickup_details(excel_path, jobcard_no):
    """
    Searches for a job card number in the Excel file and returns row data.
    - If required pickup fields (except Pick Up Driver) are missing, returns a message.
    - If Pick Up Driver starts with 'NO PICKUP' or 'EMPLOYEE PICKUP' (case-sensitive), returns a specific message.
    - If Pick Up Driver starts with 'EMPLOYEE PICKUP' and required fields are missing, returns combined message.

    Returns:
        tuple: (row_dict, status_message)
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active

    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    jobcard_no = str(jobcard_no).strip().lower()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        current_jobcard = str(row[0]).strip().lower()
        if current_jobcard == jobcard_no:
            row_dict = dict(zip(headers, row))

            # Prepare pickup driver value
            pickup_driver = row_dict.get("Pick Up Driver")
            pickup_driver_value = str(pickup_driver).strip() if pickup_driver else ""

            # Check for missing required fields (excluding pickup driver)
            required_fields = ["Pick Up Date", "Pick Up Time", "Distance in KM"]
            missing_required = any(
                row_dict.get(field) is None or str(row_dict.get(field)).strip() == ""
                for field in required_fields
            )

            # Combined check: EMPLOYEE PICKUP and missing mandatory fields
            if pickup_driver_value.startswith("EMPLOYEE PICKUP") and missing_required:
                return row_dict, 'employee pickup with some mandatory data is missing'

            # Check if only required fields are missing
            if missing_required:
                return row_dict, 'Some mandatory values except Pickup Driver is missing'

            # Check if pickup driver is missing
            if pickup_driver_value == "":
                return row_dict, 'Pickup Driver is missing'

            # Check if employee pickup
            if pickup_driver_value.startswith("EMPLOYEE PICKUP"):
                return row_dict, 'employee pickup'

            # Check if no pickup
            if pickup_driver_value.startswith("NO PICKUP"):
                return row_dict, 'no pickup'

            # Format Pick Up Date if it's a datetime
            pickup_date = row_dict.get("Pick Up Date")
            if isinstance(pickup_date, datetime):
                row_dict["Pick Up Date"] = pickup_date.strftime('%d-%m-%Y')

            return row_dict, 'fields extracted'

    return {}, None



from datetime import datetime

def get_current_date_and_time_pickup():
    now = datetime.now()
    date_str = now.strftime("%d-%m-%Y")
    time_str = now.strftime("%H:%M:%S")
    return date_str, time_str

def normalize_vehicle_number(vn):
    """
    Normalize a vehicle number to a standard comparison form.
    E.g., 'KL07YA001' -> 'KL7YA1'
    """
    match = re.match(r'([A-Z]{2})0*(\d{1,2})([A-Z]{1,2})0*(\d+)', vn.upper())
    if match:
        state, district, series, number = match.groups()
        return f"{state}{int(district)}{series}{int(number)}"
    return None

def vehicle_number_found(reference_vn, target_text):
    """
    Check if the reference vehicle number appears in the target application data,
    allowing for flexible formats.
    """
    normalized_ref = normalize_vehicle_number(reference_vn)
    
    # Find all possible vehicle-like patterns in the target text
    possible_matches = re.findall(r'[A-Z]{2}0?\d{1,2}[A-Z]{1,2}0*\d+', target_text.upper())
    
    for match in possible_matches:
        if normalize_vehicle_number(match) == normalized_ref:
            return True, reference_vn
    return False, None



# reference_number = "KL07YA001"
# target_app_text = "Some details about vehicle KL07YA01 found in system."

# result, match = vehicle_number_found(reference_number, target_app_text)
# print(result)
# print(match)
# print("Match found ✅" if result else "Match not found ❌")
