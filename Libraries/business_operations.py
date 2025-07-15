import os
from datetime import datetime
import shutil
import glob
import pandas as pd
from pathlib import Path
from robot.api  import logger
from openpyxl.styles import numbers
from openpyxl import load_workbook


def move_dms_report_to_currentdate_folder(current_date):
    try:
        # Get the current script's location
        # current_path = Path(__file__).resolve()
        # current_path = Path(r"C:\JobcardOpeningIntegrated")
        # root_folder = current_path.parents[len(current_path.parts) - 3]
        root_folder = Path(r"C:\JobcardOpeningIntegrated")
        print("Project Root Folder:", root_folder)

        parent_folder = os.path.join(root_folder, 'ProcessRelatedFolders')
        print("Process Related Folders path:", parent_folder)

        downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
        print("System Downloads Path:", downloads_folder)

        pcs_downloads_dir = os.path.join(parent_folder, current_date, 'Downloads')
        print("pcs downloads folder:", pcs_downloads_dir)

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        cons_report_file_name = f"consolidated_jobcard_report{timestamp}.xlsx"
        consolidated_report_path = os.path.join(parent_folder, current_date, 'InProgress', cons_report_file_name)
        print("consolidated_report_path:", consolidated_report_path)

        csv_files = glob.glob(os.path.join(downloads_folder, '*.csv'))

        if csv_files:
            carry_report_full_path = max(csv_files, key=os.path.getmtime)
            print("Full Path of carry ratio report from downloads dir:", carry_report_full_path)

            file_name_only = os.path.basename(carry_report_full_path)
            file_name_without_ext = os.path.splitext(file_name_only)[0]
            full_path_with_xlsx = os.path.splitext(carry_report_full_path)[0] + ".xlsx"
            print("full_path_with_xlsx:", full_path_with_xlsx)

            # Read and process CSV
            with open(carry_report_full_path, "r", encoding="ISO-8859-1") as file:
                lines = file.readlines()

            header_row = None
            for i, line in enumerate(lines):
                if "srl no" in line.lower():
                    header_row = i
                    break

            if header_row is None:
                raise Exception("Could not find the header row in the CSV file.")

            metadata = [line.strip().split(",") for line in lines[:header_row]]
            table_data = pd.read_csv(
                carry_report_full_path,
                delimiter=",",
                encoding="ISO-8859-1",
                skiprows=header_row,
                engine="python",
                on_bad_lines='skip'
            )

            table_data = table_data.apply(lambda x: x.str.replace(":", "").str.strip() if x.dtype == "object" else x)

            # Format 'Promised Dt' to string as "DD-MM-YYYY HH:MM"
            if "Promised Dt" in table_data.columns:
                table_data["Promised Dt"] = pd.to_datetime(
                    table_data["Promised Dt"], errors='coerce', dayfirst=True
                ).dt.strftime("%d-%m-%Y %H:%M")

            print("Data cleaned successfully!")

            # Write to Excel
            with pd.ExcelWriter(full_path_with_xlsx, engine="openpyxl") as writer:
                workbook = writer.book
                sheet = workbook.create_sheet(title=file_name_without_ext)

                # Write metadata row by row
                for i, row in enumerate(metadata, start=1):
                    for j, value in enumerate(row, start=1):
                        sheet.cell(row=i, column=j, value=value)

                # Write actual table data starting below metadata
                table_data.to_excel(writer, sheet_name=file_name_without_ext, index=False, startrow=len(metadata))

            file_name_with_ext_xlsx = os.path.basename(full_path_with_xlsx)
            carry_over_ratio_report_path = os.path.join(pcs_downloads_dir, file_name_with_ext_xlsx)
            print(f"carry_over_ratio_report_path: {carry_over_ratio_report_path}.")

            shutil.move(full_path_with_xlsx, pcs_downloads_dir)
            print(f"The latest Excel file has been moved to {pcs_downloads_dir}.")

            return consolidated_report_path, carry_over_ratio_report_path

        else:
            print("No CSV files found in the Downloads folder.")

    except Exception as e:
        print(f"Error while moving the report: {str(e)}")
        raise Exception(e)
# move_dms_report_to_currentdate_folder("2025-04-09")

def create_failure_report(failure_report_path):
    if os.path.exists(failure_report_path):
        return True
    else:
        print("Failure Report File does not exist. Creating")
        df_failed = pd.DataFrame(columns=["Job Card No","registration no","OMR","ENGINE NUM","Service Type Code","Service Type Description","Sub Service Type","Service Advisor Code",
                                            "Name of Service Advisor","Vehicle ID","Vehicle Model","Vehicle Variant","Color","Vehicle Sales Date",
                                            "Recall Status","Recall Code","Extended Warranty","MCP","Technical Campaign",
                                            "Promised vehicle delivery date and time","Customer ID","Customer Name","Address 1","Address 2","Address 3",
                                            "City","State","Phone","Mobile","Demand Code","Carry Over Ratio Jobcard Report Name","DMS Execution Status",
                                            "ERP Execution Status","Processed Date","Jobcard Opening From Date","Jobcard Opening To Date","Exception Reason"])
        df_failed.to_excel(failure_report_path, index=False)
        print("Filtered data saved as"+ failure_report_path)

def update_failure_report(consolidated_report_path, failure_report_path):
    # Read the consolidated_report_path Excel file
    consolidated_report_df = pd.read_excel(consolidated_report_path)
    print(consolidated_report_df)

    # Read the failure_report_path Excel file
    failure_report_df = pd.read_excel(failure_report_path)
    print(failure_report_df)

    # Filter rows where the 'Status' column is 'Fail'
    failed_rows = consolidated_report_df[consolidated_report_df["DMS Execution Status"] == "Fail"]

    # Check if the file exists
    if os.path.exists(failure_report_path):
        # Load existing data
        existing_df = pd.read_excel(failure_report_path)

        # Append new failed rows
        updated_df = pd.concat([existing_df, failed_rows], ignore_index=True)
    else:
        # If the file doesn't exist, create a new one
        updated_df = failed_rows
    
    # Save back to Excel
    updated_df.to_excel(failure_report_path, index=False)

    print("New failed rows appended to"+ failure_report_path +"successfully!")



def validate_erp_report_vs_dms_report(carry_over_ratio_report_path, erp_report_path, consolidated_report_path, failed_report_path):
    try:
        # carry_over_ratio_report_path, erp_report_path, consolidated_report_path, failed_report_path
        # carry_over_ratio_report_path = "E:/JobcardOpeningIntegrated/ProcessRelatedFolders/2025-05-05/Downloads/ELM.FOB20250505035841.xlsx"
        # erp_report_path = "E:/JobcardOpeningIntegrated/ProcessRelatedFolders/2025-05-05/Downloads/ERP_Report_20250505155521.xlsx"
        # failed_report_path = "E:/JobcardOpeningIntegrated/ProcessRelatedFolders/2025-05-05/InProgress/consolidated_failed_report.xlsx"
        # failed_report_path = ""
        # consolidated_report_path = "E:/JobcardOpeningIntegrated/ProcessRelatedFolders/2025-05-05/InProgress/consolidated_jobcard_report20250505153601.xlsx"

        logger.info(failed_report_path)
        logger.info(carry_over_ratio_report_path)
        logger.info(erp_report_path)
        logger.info(consolidated_report_path)

        # Read the first Excel file
        erp_report_df = pd.read_excel(erp_report_path)
        print(erp_report_df)

        # Read the second Excel file and skip the first 6 rows
        elm_report_df = pd.read_excel(carry_over_ratio_report_path, skiprows=6)
        print(elm_report_df)

        # Standardize column names and remove whitespace if any
        erp_report_df.columns = erp_report_df.columns.str.strip()
        elm_report_df.columns = elm_report_df.columns.str.strip()

        elm_report_df['srl no'] = pd.to_numeric(elm_report_df['srl no'], errors='coerce').fillna(0).astype(int)
        elm_report_df = elm_report_df[(elm_report_df['srl no'] != 0) & (elm_report_df['Job Card No'].str.startswith('JC'))]

        elm_report_df = elm_report_df.rename(columns={'service type':'Service Type Description', 'S.A': 'Name of Service Advisor', 'Service Type': 'Service Type Code', 'Promised Dt':'Promised vehicle delivery date and time', 'Address1': 'Address 1', 'Address2': 'Address 2', 'Address3': 'Address 3', 'Location': 'Branch'})

        # Filter the 'Job Card No' and 'Registration No' columns from the second Excel file
        columns_to_extract = ['Job Card No', 'registration no', 'ENGINE NUM', 'Service Type Description', 'Name of Service Advisor', 'Service Type Code', 'Promised vehicle delivery date and time', 'Customer Name', 'Address 1', 'Address 2', 'Address 3', 'Pin', 'Branch']
        filtered_elm_df = elm_report_df[columns_to_extract]
        print(filtered_elm_df)

        # Normalize casing and trim whitespace
        erp_report_df["Manual Jobcard No"] = erp_report_df["Manual Jobcard No"].str.upper().str.strip()
        filtered_elm_df["Job Card No"] = filtered_elm_df["Job Card No"].str.upper().str.strip()

        # Merge the dataframes on 'Job Card No'
        filtered_elm_df["Exception Reason"] = filtered_elm_df["Job Card No"].apply(lambda x: "Already Entered In ERP" if x in erp_report_df["Manual Jobcard No"].values else "")
        
        print(filtered_elm_df)

        if os.path.exists(consolidated_report_path):
            if os.path.exists(failed_report_path):
                print("failed_report_path File exists!")

                # Load the consolidated failed report
                consolidated_failed_report = pd.read_excel(failed_report_path)
                print(consolidated_failed_report)

                # Standardize column names (trim spaces)
                consolidated_failed_report.columns = consolidated_failed_report.columns.str.strip()
                filtered_elm_df.columns = filtered_elm_df.columns.str.strip()

                filtered_failed_df = pd.concat([consolidated_failed_report, filtered_elm_df])
                filtered_failed_df = filtered_failed_df.drop_duplicates(subset=['Job Card No'])
                print(filtered_failed_df)

                # Add a "Status" column with "Retry Attempted"
                filtered_failed_df.loc[:, "Exception Reason"] = "Previous Run Failed. Retry Attempted"

                filtered_consolidated_df = pd.concat([filtered_elm_df, filtered_failed_df], ignore_index=True)
                filtered_consolidated_df.to_excel(consolidated_report_path, index=False)
        else:
            print("consolidated report does not exist")
            df_summary = pd.DataFrame(columns=["Job Card No","registration no",'Branch',"OMR","ENGINE NUM","Service Type Code","Service Type Description","Sub Service Type","Service Advisor Code",
                                            "Name of Service Advisor","Vehicle ID","Vehicle Model","Vehicle Variant","Color","Vehicle Sales Date",
                                            "Recall Status","Recall Code","Extended Warranty","MCP","Technical Campaign",
                                            "Promised vehicle delivery date and time","Customer ID","Customer Name","Address 1","Address 2","Address 3","Pin",
                                            "City","State","Phone","Mobile","Demand Code","Carry Over Ratio Jobcard Report Name","DMS Execution Status",
                                            "ERP Execution Status","Processed Date","Jobcard Opening From Date","Jobcard Opening To Date","Exception Reason"])
            df_summary = pd.concat([df_summary, filtered_elm_df], ignore_index=True)
            df_summary.to_excel(consolidated_report_path, index=False)
            print("Filtered data saved as"+ consolidated_report_path)
    except Exception as e:
        print(f"Error while validating ERP report With DMS Report: {str(e)}")
        raise Exception(e)
       
# validate_erp_report_vs_dms_report()


def update_final_jobcard_details_report(consolidated_report_path, carry_over_ratio_report_path):
    try:
        # consolidated_report_path, carry_over_ratio_report_path
        # consolidated_report_path = "D:\\JobcardOpeningProcess\\ProcessRelatedFolders\\2025-03-06\\InProgress\\consolidated_jobcard_report.xlsx"
        # carry_over_ratio_report_path = "D:\\JobcardOpeningProcess\\ProcessRelatedFolders\\2025-03-06\\Downloads\\ELM.FOB20250306034352.xlsx"
        print(f"carry_over_ratio_report_path {carry_over_ratio_report_path}")
        print(f"file_path {consolidated_report_path}")
        
        # Read the source Excel file into a DataFrame
        df_source = pd.read_excel(carry_over_ratio_report_path, skiprows=6)

        # Clean column names by stripping any leading/trailing spaces
        df_source.columns = df_source.columns.str.strip()

        df_source = df_source.rename(columns={'service type':'Service Type Description', 'S.A': 'Name of Service Advisor', 'Service Type': 'Service Type Code', 'Promised Dt':'Promised vehicle delivery date and time', 'Address1': 'Address 1', 'Address2': 'Address 2', 'Address3': 'Address 3', 'Location': 'Branch'})
        
        # Now, print the columns to ensure no extra spaces
        print("Cleaned Columns in the source Excel file:")
        print(df_source.columns)

        df_source['srl no'] = pd.to_numeric(df_source['srl no'], errors='coerce').fillna(0).astype(int)
        df_source = df_source[(df_source['srl no'] != 0) & (df_source['Job Card No'].str.startswith('JC'))]

        for index, row in df_source.iterrows():
            print(f"srl no: {row['srl no']}, Job Card No: {row['Job Card No']}")    

        # Extract the 'Job Card No', 'Registration No' and 'ENGINE NUM' columns from the source DataFrame
        try:

            df_dest = df_source[['Job Card No', 'registration no', 'ENGINE NUM', 'Service Type Description', 'Name of Service Advisor', 'Service Type Code', 'Promised vehicle delivery date and time', 'Customer Name', 'Address 1', 'Address 2', 'Address 3', 'Pin', 'Branch']]
            print("Job Card No, Registration No, Service Type Description, Name of Service Advisor, Service Type Code, Promised vehicle delivery date and time, Customer Name, Address 1, Address 2, Address 3  and ENGINE NUM columns extracted successfully.")
        except KeyError as e:
            print(f"Error: {e} column not found in the source file.")
            raise Exception(e)

        if os.path.exists(consolidated_report_path):
            print("exist")
            # Read the final_jobcard_details_report Excel file into a DataFrame
            df_summary = pd.read_excel(consolidated_report_path)

            # Create a new DataFrame for the destination with the two selected columns
            new_df_filtered = df_dest[~df_dest['Job Card No'].isin(df_summary['Job Card No'])]

            # Append the filtered new data to the summary DataFrame
            combined_df = pd.concat([df_summary, new_df_filtered], ignore_index=True)

            # Write the combined DataFrame back to the summary Excel file
            combined_df.to_excel(consolidated_report_path, index=False)

        else:
            print("consolidated report does not exist")
            df_summary = pd.DataFrame(columns=["Job Card No","registration no",'Branch',"OMR","ENGINE NUM","Service Type Code","Service Type Description","Sub Service Type","Service Advisor Code",
                                            "Name of Service Advisor","Vehicle ID","Vehicle Model","Vehicle Variant","Color","Vehicle Sales Date",
                                            "Recall Status","Recall Code","Extended Warranty","MCP","Technical Campaign",
                                            "Promised vehicle delivery date and time","Customer ID","Customer Name","Address 1","Address 2","Address 3","Pin",
                                            "City","State","Phone","Mobile","Demand Code","Carry Over Ratio Jobcard Report Name","DMS Execution Status",
                                            "ERP Execution Status","Processed Date","Jobcard Opening From Date","Jobcard Opening To Date","Exception Reason"])
            df_summary = pd.concat([df_summary, df_dest], ignore_index=True)

            for index, row in df_summary.iterrows():
                print(row)

            df_summary.to_excel(consolidated_report_path, index=False)

        print("\nRows not starting with 'JC' have been removed, and the cleaned data has been saved.")
    except Exception as e:
        print(f"Error while Updating Consolidated Report: {str(e)}")
        raise Exception(e)

# update_final_jobcard_details_report()

def increase_cordinates(region, x, y, w, h):
    x = int(x.strip('"'))
    y = int(y.strip('"'))
    w = int(w.strip('"'))
    h = int(h.strip('"'))
    
    region[0] += x
    region[1] += y
    region[2] += w
    region[3] += h
    return region

def update_output_excel_with_extracted_values(extracted_data_dict, summary_file, jobcard_no, carry_over_ratio_report_path, empty_fields):
    try:
        # extracted_data_dict = {'Job Card No': 'JC24019576', 'OMR': '§5537|', 'Service Type Code': 'PMS', 'Service Type Description': 'PAID SERVICE', 'Sub Service Type': 'Pmsi10', 'Service Advisor Code': 'ELMSUPO2', 'Vehicle ID': 'MA3ETDE1S00184537', 'Vehicle Model': 'AVH310', 'Vehicle Variant': 'CEA4CL1', 'Color': 'ZNZ', 'Vehicle Sales Date': '28-02-2015', 'Extended Warranty': 'YIN) [N |', 'MCP': '', 'Technical Campaign': '', 'Promised Delivery Date Time': '(04-03-2025 09:37 |ct', 'Customer ID': '2141896951', 'Customer Name': 'JAYALEKSHMI', 'Address 1': 'LEKSHMI NARAYANA BHAVANAM', 'Address 2': 'TTHEKKEKARA', 'Address 3': 'TTHEKKEKARA', 'City': 'KOTTAYAM', 'State': 'KERALA', 'Phone': '', 'Mobile': '8547930342'}
        # jobcard_no = "JC24019576"
        # carry_over_ratio_report_path = "D:\\JobcardOpeningProcess\\ProcessRelatedFolders\\2025-03-04\\Downloads\\ELM.FOB20250304042600.xlsx"
        # summary_file = "D:\\JobcardOpeningProcess\\ProcessRelatedFolders\\2025-03-04\\InProgress\\consolidated_jobcard_report.xlsx"
        # # Read the existing Excel file into a DataFrame
        df_existing = pd.read_excel(summary_file)

        #----------------------- This block is for updating excel columns "Carry Over Ratio Jobcard Report Name", 
        # "Processed Date","Jobcard Opening From Date","Jobcard Opening To Date"-----------------------------------
        
        # Get the file name of carry_over_ratio_report_path
        carry_over_ratio_report_name = Path(carry_over_ratio_report_path).name

        # Print the file name
        print(carry_over_ratio_report_name)

        # Read the carry_over_ratio_report excel file into a DataFrame
        df_carry = pd.read_excel(carry_over_ratio_report_path,header=None)

        jc_open_from_date = df_carry.at[1, 1]
        logger.info(jc_open_from_date)
        jc_open_to_date = df_carry.at[1, 3]
        logger.info(jc_open_to_date)

        #--------------------------------------------------------------------------------

        # Get the current date in YYYY-MM-DD format
        current_date = datetime.now().strftime("%Y-%m-%d")

        logger.info(df_existing)

        if jobcard_no in df_existing["Job Card No"].values:
            for index, row in df_existing.iterrows():
                if row["Job Card No"] == jobcard_no:
                    # Update each column based on the dictionary
                    df_existing.at[index, 'OMR'] = extracted_data_dict['OMR']
                    # df_existing.at[index, 'Service Type Code'] = extracted_data_dict['Service Type Code']
                    # df_existing.at[index, 'Service Type Description'] = extracted_data_dict['Service Type Description']
                    df_existing.at[index, 'Sub Service Type'] = extracted_data_dict['Sub Service Type']
                    df_existing.at[index, 'Service Advisor Code'] = extracted_data_dict['Service Advisor Code']
                    # df_existing.at[index, 'Name of Service Advisor'] = extracted_data_dict['Name of Service Advisor']
                    df_existing.at[index, 'Vehicle ID'] = extracted_data_dict['Vehicle ID']
                    df_existing.at[index, 'Vehicle Model'] = extracted_data_dict['Vehicle Model']
                    df_existing.at[index, 'Vehicle Variant'] = extracted_data_dict['Vehicle Variant']
                    df_existing.at[index, 'Color'] = extracted_data_dict['Color']
                    df_existing.at[index, 'Vehicle Sales Date'] = extracted_data_dict['Vehicle Sales Date']
                    df_existing.at[index, 'Extended Warranty'] = extracted_data_dict['Extended Warranty']
                    df_existing.at[index, 'MCP'] = extracted_data_dict['MCP']
                    df_existing.at[index, 'Technical Campaign'] = extracted_data_dict['Technical Campaign']
                    # df_existing.at[index, 'Promised vehicle delivery date and time'] = extracted_data_dict['Promised Delivery Date Time']
                    df_existing.at[index, 'Customer ID'] = extracted_data_dict['Customer ID']
                    # df_existing.at[index, 'Customer Name'] = extracted_data_dict['Customer Name']
                    # df_existing.at[index, 'Address 1'] = extracted_data_dict['Address 1']
                    # df_existing.at[index, 'Address 2'] = extracted_data_dict['Address 2']
                    # df_existing.at[index, 'Address 3'] = extracted_data_dict['Address 3']
                    df_existing.at[index, 'City'] = extracted_data_dict['City']
                    df_existing.at[index, 'State'] = extracted_data_dict['State']
                    df_existing.at[index, 'Phone'] = extracted_data_dict['Phone']
                    df_existing.at[index, 'Mobile'] = extracted_data_dict['Mobile']
                    # df_existing.at[index, 'Mobile'] = extracted_data_dict[jobcard_number].get('Mobile', '')
                    df_existing.at[index, 'Carry Over Ratio Jobcard Report Name'] = carry_over_ratio_report_name
                    df_existing.at[index, 'Processed Date'] = current_date
                    df_existing.at[index, 'Jobcard Opening From Date'] = jc_open_from_date
                    df_existing.at[index, 'Jobcard Opening To Date'] = jc_open_to_date
                    if empty_fields != "":
                        logger.info(empty_fields)
                        df_existing.at[index, 'DMS Execution Status'] = "Fail"
                        df_existing.at[index, 'Exception Reason'] = empty_fields
            # Save the updated DataFrame back to Excel
            df_existing.to_excel(summary_file, index=False)

        print("Extracted datas added successfully to the excel!")
    except Exception as e:
        print(f"Error while Updating Consolidated Report: {str(e)}")
        raise Exception(e)
# update_output_excel_with_extracted_values()

def update_demand_codes_in_output_excel(summary_file, job_card_no, demand_code):
    try:
        # Read the existing Excel file into a DataFrame
        df = pd.read_excel(summary_file)

        # job_card_no is the actual Job Card No 
        # demand_code in the form of "AD1236,VY1253". The value to set for Demand Code in the excel.

        # Update the Demand Code column for the specific Job Card No
        df.loc[df['Job Card No'] == job_card_no, 'Demand Code'] = demand_code

        # Save the updated dataframe back to the Excel file (optional)
        df.to_excel(summary_file, index=False)
        return True
    except Exception as e:
        print(f"Error while updating Demand Code in the report: {str(e)}")
        return False 


# def update_execution_status_in_summary_report(summary_file, job_card_no, status, column_name):
#     # Read the existing Excel file into a DataFrame
#     df = pd.read_excel(summary_file)

#     # In case of DMS, column_name is 'DMS Execution Status'. In case of ERP column_name is 'ERP Execution Status'
#     df.loc[df['Job Card No'] == job_card_no, column_name] = status

#     # Save the updated dataframe back to the Excel file (optional)
#     df.to_excel(summary_file, index=False) 

# def update_execution_status_in_summary_report(summary_file, job_card_no, status, column_name):

#     logger.info(status)
#     # Clean up the status string
#     # cleaned_status = str(status).strip().replace('\n', ' ').replace('\r', '')
#     cleaned_status = str(status).replace('\r', '').replace('\n', ' ').strip()

#     logger.info(cleaned_status)
#     # Read the existing Excel file into a DataFrame
#     df = pd.read_excel(summary_file)

#     # Update the appropriate cell
#     df.loc[df['Job Card No'] == job_card_no, column_name] = cleaned_status

#     # Save the updated dataframe back to the Excel file
#     df.to_excel(summary_file, index=False)

# import pandas as pd
# import logging

# logger = logging.getLogger(__name__)

# def update_execution_status_in_summary_report(summary_file, job_card_no, status, column_name):
#     logger.info(status)
#     cleaned_status = str(status).replace('\r', '').replace('\n', ' ').strip()
#     logger.info(cleaned_status)

#     # Read the existing Excel file into a DataFrame
#     df = pd.read_excel(summary_file)

#     # Find the row index for the given job_card_no
#     row_index = df.index[df['Job Card No'] == job_card_no].tolist()

#     if not row_index:
#         logger.warning(f"Job Card No '{job_card_no}' not found in the summary report.")
#         return

#     idx = row_index[0]

#     # Get the existing value in the cell
#     existing_value = str(df.at[idx, column_name])
#     if existing_value.lower() == 'nan':
#         existing_value = ''

#     # Add line break before appending if needed
#     updated_value = f"{existing_value}\n{cleaned_status}" if existing_value else cleaned_status
#     df.at[idx, column_name] = updated_value

#     # Save with ExcelWriter to preserve formatting
#     with pd.ExcelWriter(summary_file, engine='openpyxl') as writer:
#         df.to_excel(writer, index=False)


import pandas as pd
import logging

logger = logging.getLogger(__name__)

def update_execution_status_in_summary_report(summary_file, job_card_no, status, column_name):
    logger.info(status)
    cleaned_status = str(status).replace('\r', '').replace('\n', ' ').strip()
    logger.info(cleaned_status)

    # Read the existing Excel file into a DataFrame
    df = pd.read_excel(summary_file)

    # Find the row index for the given job_card_no
    row_index = df.index[df['Job Card No'] == job_card_no].tolist()

    if not row_index:
        logger.warning(f"Job Card No '{job_card_no}' not found in the summary report.")
        return

    idx = row_index[0]

    if cleaned_status.lower() in ['success', 'fail']:
        # Overwrite if status is Success or Fail
        df.at[idx, column_name] = cleaned_status
    else:
        # Get the existing value in the cell
        existing_value = str(df.at[idx, column_name])
        if existing_value.lower() == 'nan':
            existing_value = ''

        # Append with line break if not empty
        updated_value = f"{existing_value}\n{cleaned_status}" if existing_value else cleaned_status
        df.at[idx, column_name] = updated_value

    # Save with ExcelWriter to preserve formatting
    with pd.ExcelWriter(summary_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

def update_execution_status_as_null(summary_file, job_card_no, status, column_name):
    logger.info(status)

    # Read the existing Excel file into a DataFrame
    df = pd.read_excel(summary_file)

    # Find the row index for the given job_card_no
    row_index = df.index[df['Job Card No'] == job_card_no].tolist()

    if not row_index:
        logger.warning(f"Job Card No '{job_card_no}' not found in the summary report.")
        return

    idx = row_index[0]
    df.at[idx, column_name] = ''
    
    # Save with ExcelWriter to preserve formatting
    with pd.ExcelWriter(summary_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)



def update_recall_in_output_excel(summary_file, job_card_no, recall_code, recall_status):
    try:
        # Read the existing Excel file into a DataFrame
        df = pd.read_excel(summary_file)

        # job_card_no is the actual Job Card No 
        # recall_code in the form of "Cir.#: D-48A/2023". The value updated in the Recall Code column in the excel.

        # Update the Recall Code column for the specific Job Card No
        df.loc[df['Job Card No'] == job_card_no, 'Recall Code'] = recall_code
        
        # Update the Recall Status column for the specific Job Card No
        df.loc[df['Job Card No'] == job_card_no, 'Recall Status'] = recall_status

        # Save the updated dataframe back to the Excel file (optional)
        df.to_excel(summary_file, index=False)
        return True
    except Exception as e:
        print(f"Error while updating Recall Code in the report: {str(e)}")
        return False

def get_current_date_fun():
    # Get the current date in YYYY-MM-DD format
    current_date = datetime.now().strftime("%Y-%m-%d")
    # Get today's date in the required format (e.g., 20250123)
    today_date = datetime.today().strftime('%Y%m%d')

    # Get current timestamp in the required format
    date_timestamp = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    return  current_date, date_timestamp

# def get_process_root_directory():
#     # ----------------- Getting the root folder path ----------------
#     # Get the absolute path of the current script
#     current_file_path = Path(__file__).resolve()

#     # Move up one level to reach "JobcardOpeningProcess"
#     project_root = current_file_path.parents[1]

#     print(f"Project Root Directory: {project_root}")
#     #---------------------------------------------------------------
#     return project_root
    # from pathlib import Path

def get_process_root_directory():
    # Return a fixed path for the project root
    project_root = Path(r"C:\JobcardOpeningIntegrated")
    print(f"Project Root Directory: {project_root}")
    return project_root


def process_related_folder_creation(curr_date):
    try:
        #----------------- Getting the root folder path ----------------
        project_root = get_process_root_directory()
        #---------------------------------------------------------------

        #-----Creating Screenshot folder inside root folder -----
        # Define the new folder path inside the root folder
        Screenshot_folder = project_root / "Screenshot"

        # Create the folder if it doesn't exist
        Screenshot_folder.mkdir(parents=True, exist_ok=True)
        print(f"Project Root Directory: {Screenshot_folder}")
        #---------------------------------------------------------------------


        #-----Creating DMS_captured folder inside root folder -----
        # Define the new folder path inside the root folder
        DMS_captured_folder = project_root / "DMS_captured"

        # Create the folder if it doesn't exist
        DMS_captured_folder.mkdir(parents=True, exist_ok=True)
        print(f"Project Root Directory: {DMS_captured_folder}")
        #---------------------------------------------------------------------

        #-----Creating results folder inside root folder -----
        # Define the new folder path inside the root folder
        results_folder = project_root / "Results"

        # Create the folder if it doesn't exist
        results_folder.mkdir(parents=True, exist_ok=True)
        print(f"Project Root Directory: {results_folder}")
        #---------------------------------------------------------------------

        #--------------- Creating current date folder inside "Results folder"--------
        # Define the new folder path with today's date
        results_date_folder = results_folder / curr_date

        # Create the folders if they don't exist
        results_date_folder.mkdir(parents=True, exist_ok=True)

        print(f"Folder created (or already exists) at: {results_date_folder}")
        #------------------------------------------------------------------------

        #--------------- Creating Downloads folder inside "Results/current date folder"--------
        results_downloads_folder = results_date_folder / "Downloads"

        # Create the folders if they don't exist
        results_downloads_folder.mkdir(parents=True, exist_ok=True)

        print(f"Folder created (or already exists) at: {results_downloads_folder}")
        #------------------------------------------------------------------------


        #-----Creating process_related_folders folder inside root folder -----
        # Define the new folder path inside the root folder
        process_related_folder = project_root / "ProcessRelatedFolders"

        # Create the folder if it doesn't exist
        process_related_folder.mkdir(parents=True, exist_ok=True)
        print(f"Project Root Directory: {process_related_folder}")
        #---------------------------------------------------------------------

        #--------------- Creating current date folder inside "process_related_folders"--------
        # Define the new folder path with today's date
        date_folder = process_related_folder / curr_date

        # Create the folders if they don't exist
        date_folder.mkdir(parents=True, exist_ok=True)

        print(f"Folder created (or already exists) at: {date_folder}")
        #------------------------------------------------------------------------

        # ------- "Downloads", "InProgress", "ProcessedDMS", "ProcessedERP" inside current date folder -----
        # Define the subfolders to create inside the date folder
        subfolders = ["Downloads", "InProgress", "ProcessedDMS", "ProcessedERP"]

        # Create the folders if they don't exist
        for subfolder in subfolders:
            (date_folder / subfolder).mkdir(parents=True, exist_ok=True)

        print(f"Folders created inside: {date_folder}")
        #--------------------------------------------------------
    except Exception as e:
        print(f"Error while creating the folders: {str(e)}")
        raise Exception(e)
    

def prepare_file_name_of_downloaded_erp_report(service_type):
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    print(timestamp)
    file_name = ""
    if service_type == "Regular":
        file_name = "ERP_Report_Regular_Service"+timestamp
    else:
        file_name = "ERP_Report_Bodyshop_Service"+timestamp   
    return  file_name

def process_file(file_path, columns_mapping):
    df = pd.read_excel(file_path, skiprows=3)  # Skip header rows
    df = df[list(columns_mapping.keys())]  # Select required columns
    df.rename(columns=columns_mapping, inplace=True)  # Rename columns
    return df

def combine_erp_regular_and_bodyshop_report(erp_regular_report_fullpath, erp_bodyshop_report_fullpath, downloads_folder, region_mapping_sheet, location):
    try:
        # erp_regular_report_fullpath, erp_bodyshop_report_fullpath, downloads_folder, region_mapping_sheet, location
        # erp_regular_report_fullpath="E:/JobcardOpeningIntegrated/ProcessRelatedFolders/2025-05-08/Downloads/ERP_Report_Regular_Service20250508152312.xlsx"
        # erp_bodyshop_report_fullpath="E:/JobcardOpeningIntegrated/ProcessRelatedFolders/2025-05-08/Downloads/ERP_Report_Bodyshop_Service20250508152345.xlsx"
        # downloads_folder="E:/JobcardOpeningIntegrated/ProcessRelatedFolders/2025-05-08/Downloads"
        # region_mapping_sheet="E:/JobcardOpeningIntegrated/Mapping/Location Mapping DMS ERP.xlsx"
        # location="ELAMAKKARA-SRV"
        columns_mapping = {
            "SNo": "SNo",
            "Branch": "Branch",
            "Chassis Number": "Chassis Number",
            "Registration No": "Registration No",
            "Manual Job Card No": "Manual Jobcard No"
        }

        columns_mapping2 = {
            "SNo": "SNo",
            "Branch": "Branch",
            "Chassis Number": "Chassis Number",
            "Registration Number": "Registration No",
            "Manual Jobcard No": "Manual Jobcard No"
        }

        bodyshop_df = process_file(os.path.join(downloads_folder, erp_bodyshop_report_fullpath), columns_mapping)
        
        # columns_mapping["Registration Number"] = "Registration No"  # Adjust column name for second file
        regular_df = process_file(os.path.join(downloads_folder, erp_regular_report_fullpath), columns_mapping2)

        print(regular_df.columns)
        print(bodyshop_df.columns)
        
        merged_df = pd.concat([bodyshop_df, regular_df], ignore_index=True)
        merged_df.dropna(subset=["Branch", "Chassis Number", "Registration No", "Manual Jobcard No"], how='all', inplace=True)  # Remove empty rows after merging
        
        if "SNo" in merged_df.columns:
            # Drop existing SNo column if present    
            merged_df.drop(columns=["SNo"], inplace=True)  
        # Auto-generate SNo
        merged_df.insert(0, "SNo", range(1, len(merged_df) + 1)) 
        
        # Create a timestamped filename
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        
        output_filename = f"ERP_Report_{timestamp}.xlsx"
        output_path = os.path.join(downloads_folder, output_filename)

        # ------------------------------------------------------------------------------
        # --------  reading region_mapping_sheet to get the DMS Location description----
        region_df = pd.read_excel(region_mapping_sheet, dtype=str)

        region_df.columns = region_df.columns.str.strip()

        # Get ERP location code from Location Mapping DMS ERP using DMS Location description
        erp_code_row = region_df[region_df['DMS Location description'] == location]
        
        erp_code = erp_code_row['ERP location code'].values[0]

        #Filter merged_df using this ERP location code
        filtered_df = merged_df[merged_df['Branch'] == erp_code]

        # Save the merged file
        filtered_df.to_excel(output_path, index=False)

        print(f"Merged file saved as {output_path}")

        return output_path
    except Exception as e:
        print(f"Error while merging the erp reports: {str(e)}")
        raise Exception(e)
    
# combine_erp_regular_and_bodyshop_report()

def move_report_to_destination_folder(report_path, destination_path):
    # Ensure destination folder exists
    os.makedirs(destination_path, exist_ok=True)

    # copy the file
    shutil.copy(report_path, destination_path)

    print(f"File copied to {destination_path} successfully!")

def copy_report_to_destination_folder(consolid_report_path_in_processed_dms_dir, destination_folder):
    # Ensure destination folder exists
    os.makedirs(destination_folder, exist_ok=True)

    consolidated_report_file_to_check = os.path.join(destination_folder, "consolidated_jobcard_report.xlsx")

    # Check if the file exists in the destination folder
    if not os.path.exists(consolidated_report_file_to_check):
        shutil.copy(consolid_report_path_in_processed_dms_dir, consolidated_report_file_to_check)
        print(f" {consolid_report_path_in_processed_dms_dir} File copied to {consolidated_report_file_to_check}")
    else:
        print(f"{consolid_report_path_in_processed_dms_dir} File already exists: {consolidated_report_file_to_check}")
        # Need to merge new rows from the source consolidated report with timestamp to existing consolidated report without stamp in the Results/current_date folder

        # Read existing file
        df_existing = pd.read_excel(consolidated_report_file_to_check)
        df_new = pd.read_excel(consolid_report_path_in_processed_dms_dir)

        # Append new data to existing data
        df_combined = pd.concat([df_existing, df_new], ignore_index=True)

        # Save back to consolidated file
        df_combined.to_excel(consolidated_report_file_to_check, index=False)
        print("New JC rows appended to existing consolidated_jobcard_report.xlsx")

    return  consolidated_report_file_to_check

def check_jobcard_status(file_path, job_card_no, recall_status):
    df = pd.read_excel(file_path)
    if recall_status == "No":
        match = df[
            (df['Job Card No'].astype(str) == str(job_card_no)) &
            (df['Recall Status'].str.strip().str.upper() == "NO") &
            (df['DMS Execution Status'].str.strip().str.upper() == "SUCCESS")
        ]
        return not match.empty
    else:
        match = df[
            (df['Job Card No'].astype(str) == str(job_card_no)) &
            (df['Recall Status'].str.strip().str.upper() == "YES") &
            (df['DMS Execution Status'].str.strip().str.upper() == "SUCCESS")
        ]
        return not match.empty


# abc= check_jobcard_status("E:\\JobcardOpeningIntegrated\\ProcessRelatedFolders\\2025-04-21\\InProgress\\consolidated_jobcard_report20250418221818.xlsx","JC25000911")
# print(abc)


